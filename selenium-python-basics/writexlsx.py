from openpyxl import Workbook,load_workbook
wb=Workbook()
a=load_workbook()
ws=wb.active
l=[["name","city"],["harish","delhi"],["naman","mumbai"]]
for i in l:
    ws.append(i)
wb.save("program")