from openpyxl import Workbook
from faker import Faker
wb=Workbook()
ws=wb.active
fake=Faker()
for i in range(1,11):
    ws.cell(row=i,column=1).value=fake.name()
for i in range(1,11):
    ws.cell(row=i,column=2).value=fake.email()
for i in range(1,11):
    ws.cell(row=i,column=3).value=fake.address()
wb.save("10name")
